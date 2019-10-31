# -*- coding: utf-8 -*-

from odoo import models, fields, api
from odoo.exceptions import ValidationError

import os
directory = os.path.dirname(__file__)

import base64
try:
    from openpyxl import load_workbook
except ImportError:
    pass
import io


class TvaDeclaration(models.Model):
    _name = "tva.declaration"
    _description = "Declaration Tva"
    _inherit = ['mail.thread']

    @api.one
    @api.constrains('period_id')
    def _check_period_unicity(self):
        tva_ids = self.search([('period_id', '=', self.period_id.id), ('id', '!=', self.id)])
        if tva_ids:
            raise ValidationError(u"Une autre déclaration TVA pour la même période existe déjà!")

    @api.one
    @api.constrains('state')
    def _check_closed_tva(self):
        tva_ids = self.search([('state', '=', 'draft'), ('id', '!=', self.id)])
        if tva_ids:
            raise ValidationError(u"Toutes les autres déclarations de TVA doivent être cloturées!")

    @api.depends('period_id')
    @api.multi
    def get_name(self):
        for record in self:
            name = u'Déclaration de TVA: '
            if self.period_id:
                name += record.period_id.name
            record.name = name

    name = fields.Char('Description', compute='get_name', store=True)
    period_id = fields.Many2one('date.range',
                               string=u'Période', domain=[('type_id.fiscal_period','=',True)], required=True)
    company_id = fields.Many2one('res.company',u'Société', default=lambda self: self.env.user.company_id,
                                 required=False)
    line_ids = fields.One2many(comodel_name='tva.declaration.line', inverse_name='tva_declaration_id',
                                  string=u'Décaissements', domain=[('type', '=', False)])
    encaissement_ids = fields.One2many(comodel_name='tva.declaration.line', inverse_name='tva_declaration_id',
                                  string=u'Encaissements', domain=[('type', '=', True)])

    receivable_tva_ids = fields.One2many(comodel_name='tva.line', inverse_name='tva_declaration_id',
                               string=u'TVA fournisseur', domain=[('type', '=', 'p')], compute="generate_tva_lines")
    payable_tva_ids = fields.One2many(comodel_name='tva.line', inverse_name='tva_declaration_id',
                                         string=u'TVA Client', domain=[('type', '=', 'r')], compute="generate_tva_lines")

    state = fields.Selection([
        ('draft', u'Brouillon'),
        ('done', u'Valide')], default='draft', string='Etat', readonly=True, track_visibility='onchange')

    tva_report_file = fields.Binary('Fichier Excel', attachment=True)
    name_file_excel = fields.Char(string=u"Nom fichier", required=False, )
    payment_proof_file = fields.Binary(u'Reçu de paiement', attachment=True)
    regime = fields.Char(u'Régime', compute="get_data_tva")
    annee = fields.Char(u'Année', compute="get_data_tva")
    period = fields.Char(u'Période', compute="get_data_tva")
    tax_account_id = fields.Many2one('account.account', string=u"Compte Crédit de TVA",
                                     default=lambda self: self.env.user.company_id.tax_account_id)


    payed_tax_account_id = fields.Many2one('account.account', string=u"Compte TVA due",
                                    default=lambda self: self.env.user.company_id.payed_tax_account_id)

    sum_receivable = fields.Float('Total TVA client',compute="get_tva_amount")
    sum_payable = fields.Float('Total TVA fournisseur',compute="get_tva_amount")
    sum_credit = fields.Float(u'Crédit de la période précédente')
    to_be_payed = fields.Float(u'TVA à déclarer', compute="get_tva_amount")
    move_id = fields.Many2one('account.move',u'Pièce comptable',readonly=True)

    date = fields.Date('Date',required=True)

    code_tva_202 = fields.Float(u'Réduction de 15% du crédit de la période ((190 - 170) - 130) × 15%')
    code_tva_203 = fields.Float(u'Crédit restant cumulé après réduction de 15% (201- 202)')
    code_tva_157 = fields.Float(u"Déduction prévue par l'article 125-VII du CGI (1/5 du montant de la TVA payée au cours du mois de décembre 2013)")


    @api.multi
    def get_tva_credit(self):
        for record in self:
            account_mv_line_obj = self.env['account.move.line']
            sum_credit = 0
            moves = account_mv_line_obj.search([('account_id', '=', record.tax_account_id.id),
                                                ('full_reconcile_id', '=', False)])
            for mv in moves:
                sum_credit += mv.debit - mv.credit
            if sum_credit < 0:
                sum_credit = 0
            record.sum_credit = sum_credit

    @api.multi
    def get_tva_amount(self):
        for record in self:
            sum_receivable = 0
            sum_payable = 0
            for line in record.line_ids:
                if not line.to_be_delayed:
                    sum_payable += line.amount_tva
            for line in record.encaissement_ids:
                if not line.to_be_delayed:
                    sum_receivable += line.amount_tva
            record.sum_payable = sum_payable
            record.sum_receivable = sum_receivable
            record.to_be_payed = sum_receivable - sum_payable - record.sum_credit

    # Generate Accounting Move
    @api.multi
    def genetare_tax_account_move(self):
        account_move_obj = self.env['account.move']
        for record in self:
            journal_id = self.env.user.company_id.tax_journal_id.id
            move_data = []
            record.move_id.unlink()
            # Décaissements
            for line in record.line_ids:
                if not line.to_be_delayed:
                    if line.amount_tva > 0.0:
                        move_data.append({
                            'name': record.name,
                            'date': record.date,
                            'journal_id': journal_id,
                            'account_id': line.tax_id.account_id.id,
                            'credit': line.amount_tva,
                            'invoice_id': line.invoice_tax_id.invoice_id.id,
                        })
                    if line.amount_tva < 0.0:
                        move_data.append({
                            'name': record.name,
                            'date': record.date,
                            'journal_id': journal_id,
                            'account_id': line.tax_id.account_id.id,
                            'debit': abs(line.amount_tva),
                            'invoice_id': line.invoice_tax_id.invoice_id.id,
                        })
            # Encaissements
            for line in record.encaissement_ids:
                if not line.to_be_delayed:
                    if line.amount_tva > 0.0:
                        move_data.append({
                            'name': record.name,
                            'journal_id': journal_id,
                            'date': record.date,
                            'account_id': line.tax_id.account_id.id,
                            'debit': line.amount_tva,
                            'invoice_id': line.invoice_tax_id.invoice_id.id,
                        })
                    if line.amount_tva < 0.0:
                        move_data.append({
                            'name': record.name,
                            'journal_id': journal_id,
                            'date': record.date,
                            'account_id': line.tax_id.account_id.id,
                            'credit': abs(line.amount_tva),
                            'invoice_id': line.invoice_tax_id.invoice_id.id,
                        })
            # En cas de crédit de TVA
            if record.to_be_payed < 0:
                move_data.append({
                    'name': record.name,
                    'date': record.date,
                    'journal_id': journal_id,
                    'account_id': record.tax_account_id.id,
                    'debit': abs(record.to_be_payed),
                })
                if record.sum_credit:
                    move_data.append({
                        'name': record.name,
                        'journal_id': journal_id,
                        'date': record.date,
                        'account_id': record.tax_account_id.id,
                        'credit': abs(record.sum_credit),
                    })
            # En cas de TVA due
            elif record.to_be_payed > 0:
                move_data.append({
                    'name': record.name,
                    'date': record.date,
                    'journal_id': journal_id,
                    'account_id': record.payed_tax_account_id.id,
                    'credit': record.to_be_payed,
                    'invoice_id': line.invoice_tax_id.invoice_id.id,
                })
                if record.sum_credit:
                    move_data.append({
                        'name': record.name,
                        'date': record.date,
                        'journal_id': journal_id,
                        'account_id': record.tax_account_id.id,
                        'credit': record.sum_credit,
                        'invoice_id': line.invoice_tax_id.invoice_id.id,
                    })

            line_ids = [(0, 0,mv) for mv in move_data]
            move_id = account_move_obj.create({
                'journal_id': journal_id,
                'date': record.date,
                'ref': record.name,
                'line_ids': line_ids,
            })
            record.move_id = move_id

    # Compute regime & period
    @api.multi
    @api.depends('period_id')
    def get_data_tva(self):
        for record in self:
            if record.period_id and record.period_id.fiscal_year_id:
                if len(record.period_id.fiscal_year_id.period_ids) == 12:
                    record.regime = "1"
                else:
                    record.regime = "2"
            if record.period_id:
                split_date = record.period_id.date_start
                record.annee = split_date.year
                if record.regime == "2":
                    record.period = str(((split_date.month - 1) // 3) + 1)
                else:
                    if split_date.month in ('10', '11', '12'):
                        record.period = split_date.month
                    else:
                        record.period = split_date.month

    @api.multi
    def generate_tva_file(self):
        tax_report_cells = self.env['account.tax.repport'].search([])
        tva_line_obj = self.env['tva.line']
        report_template = self.env['ir.config_parameter'].sudo().get_param('tva_encaissement_maroc.tva_rapport_id')
        if not report_template:
            raise ValidationError(u'Merci de renseigner le fichier de la déclaration au niveau de la configuration')
        for record in self:
            file = base64.b64decode(report_template)
            xls_filelike = io.BytesIO(file)
            wb = load_workbook(xls_filelike)
            ws = wb.get_active_sheet()
            sum_sale_ht = sum(l.amount_ht for l in record.encaissement_ids.filtered(lambda r: r.to_be_delayed is False))
            for cell in tax_report_cells:
                if cell.code:
                    line = tva_line_obj.search([('tva_declaration_id', '=', record.id), ('tax_id.code', '=', cell.code)])
                    if cell.type == 'tax':
                        ws[cell.cell].value = line.amount
                    if cell.type == 'base':
                        ws[cell.cell].value = line.amount_ht
                if cell.code == 202:
                    ws[cell.cell].value = record.code_tva_202
                if cell.code == 203:
                    ws[cell.cell].value = record.code_tva_203
                if cell.code == 157:
                    ws[cell.cell].value = record.code_tva_157
                if cell.code == 10:
                    ws[cell.cell].value = sum_sale_ht
                if cell.code == 170:
                    ws[cell.cell].value = record.sum_credit
            wb.save(os.path.join(directory, 'tva.xlsx'))
            tva_report_file = base64.encodestring(open(os.path.join(directory, 'tva.xlsx'), 'rb').read())
            extension = 'xlsx'
            filename = u"Déclaration TVA-"+record.env.user.company_id.name+'-'+record.period+'-'+record.annee
            name = "%s.%s" % (filename, extension)
            record.write({'tva_report_file': tva_report_file, 'name_file_excel': name})

    @api.multi
    def action_draft(self):
        for record in self:
            record.state = 'draft'

    @api.multi
    def validate(self):
        for record in self:
            if not record.payment_proof_file:
                raise ValidationError(u"Merci d'attacher le reçu de paiement!")
            record.state = 'done'

    # Get Payment Type
    def get_payment(self, move):
        paiement_type = False
        for move_line in move.line_ids:
            if move_line.payement_method_id:
                paiement_type = move_line.payement_method_id.id
                break
        return paiement_type

    # Compute receivable_tva_ids & payable_tva_ids
    @api.multi
    def generate_tva_lines(self):
        tva_obj = self.env['tva.line']
        receivable_tva_ids = tva_obj
        payable_tva_ids = tva_obj
        for record in self:
            tva_obj.search([('tva_declaration_id', '=', record.id)]).unlink()
            data = {}
            for line in record.line_ids:
                if not line.to_be_delayed:
                    if data.get(line.tax_id.id, False):
                        data[line.tax_id.id]['amount'] += line.amount_tva
                        data[line.tax_id.id]['amount_ht'] += line.amount_ht
                    else:
                        data[line.tax_id.id] = {
                            'tva_declaration_id': record.id,
                            'tax_id': line.tax_id.id,
                            'amount': line.amount_tva,
                            'amount_ht': line.amount_ht,
                            'type': 'r'
                        }
            for line in record.encaissement_ids:
                if not line.to_be_delayed:
                    if data.get(line.tax_id.id, False):
                        data[line.tax_id.id]['amount'] += line.amount_tva
                        data[line.tax_id.id]['amount_ht'] += line.amount_ht
                    else:
                        data[line.tax_id.id] = {
                            'tva_declaration_id': record.id,
                            'tax_id': line.tax_id.id,
                            'amount': line.amount_tva,
                            'amount_ht': line.amount_ht,
                            'type': 'p'
                        }
            for data_tva in data:
                if data[data_tva]['type'] == 'r':
                    receivable_tva_ids += tva_obj.create(data[data_tva])
                if data[data_tva]['type'] == 'p':
                    payable_tva_ids += tva_obj.create(data[data_tva])
            self.receivable_tva_ids = receivable_tva_ids
            self.payable_tva_ids = payable_tva_ids

    # Get Payment Counterpart
    def generate_liquidity_moves(self, move):
        move_pay_ids = []
        for move_line in move.line_ids:
            if move_line.debit > 0:
                reconcial_ids = self.env['account.partial.reconcile'].search(
                    [('debit_move_id', '=', move_line.id)])
                if reconcial_ids:
                    for rec in reconcial_ids:
                        move_pay_ids.append([rec.credit_move_id, rec.amount])
            if move_line.credit > 0:
                reconcial_ids = self.env['account.partial.reconcile'].search(
                    [('credit_move_id', '=', move_line.id)])
                if reconcial_ids:
                    for rec in reconcial_ids:
                        move_pay_ids.append([rec.debit_move_id, rec.amount])
        return move_pay_ids

    @api.multi
    def get_payment_moves(self):
        move_obj = self.env['account.move']
        domain = [('date', '>=', self.period_id.date_start),
                  ('date', '<=', self.period_id.date_end),
                  ('journal_id.type', 'in', ('cash', 'bank'))]
        move_ids = move_obj.search(domain)
        print('moooooooooooooo',move_ids)
        return move_ids

    # Generate Data Lines
    @api.multi
    def generate_data(self):
        move_obj = self.env['account.move']
        tva_declaration_line_obj = self.env['tva.declaration.line']
        for record in self:
            record.line_ids.filtered(lambda r: r.delayed is False).unlink()
            record.encaissement_ids.filtered(lambda r: r.delayed is False).unlink()
            lines_delayed = tva_declaration_line_obj.search([('to_be_delayed', '=', 'True'),
                                                             ('tva_declaration_id', '!=', record.id)])

            lines_delayed.write({
                                 'tva_declaration_id': record.id,
                                 'to_be_delayed': False,
                                 'delayed': True,
                                })
            move_ids = self.get_payment_moves()
            seq = 0
            for move in move_ids:
                if move.journal_id.default_credit_account_id.user_type_id.type == 'liquidity':
                    move_pay_ids = self.generate_liquidity_moves(move)
                    pay_method = self.get_payment(move)
                    for move_pay_line in move_pay_ids:
                        move_pay = move_pay_line[0]
                        if move_pay.invoice_id and move_pay.invoice_id.invoice_line_ids:
                            tax_dict = {}
                            total_amount = move_pay_line[1]
                            num_taxes = len(move_pay.invoice_id.tax_line_ids)
                            for tax_line in move_pay.invoice_id.tax_line_ids:
                                if total_amount == move_pay.invoice_id.amount_total:
                                    amount_tax = tax_line.amount
                                    amount_ht = tax_line.base
                                    amount_ttc = amount_ht+amount_tax
                                    tax_dict[tax_line.tax_id.id] = [amount_ht, amount_tax, amount_ttc]
                                else:
                                    coef = tax_line.tax_id.amount/(100+tax_line.tax_id.amount)
                                    amount_tax = (total_amount*coef/num_taxes)
                                    amount_ht = amount_tax/(tax_line.tax_id.amount*0.01 or 1)
                                    amount_ttc = amount_ht+amount_tax
                                    tax_dict[tax_line.tax_id.id] = [amount_ht, amount_tax, amount_ttc]
                            if move_pay.invoice_id.type in ('out_refund', 'in_refund'):
                                tax_dict[tax_line.tax_id.id] = [-amount_ht, -amount_tax, -amount_ttc]
                            type = False
                            if move_pay.invoice_id.type in ('out_invoice', 'out_refund'):
                                type = True
                            for tax_line in move_pay.invoice_id.tax_line_ids:
                                invoice_number = move_pay.invoice_id.number
                                if not type:
                                    seq+=1
                                    invoice_number = move_pay.invoice_id.reference or move_pay.invoice_id.number
                                s = seq
                                if type:
                                    s=0
                                data = {
                                    'sequence': s,
                                    'tva_declaration_id': record.id,
                                    'invoice_id': move_pay.invoice_id.id,
                                    'invoice_number': invoice_number,
                                    'partner_id': move_pay.invoice_id.partner_id.id,
                                    'id_fisc': move_pay.invoice_id.partner_id.id_fisc,
                                    'partner_name': move_pay.invoice_id.partner_id.name,
                                    'ice': move_pay.invoice_id.partner_id.ice,
                                    'description': move_pay.invoice_id.invoice_line_ids[0].name,
                                    'amount_ht': tax_dict[tax_line.tax_id.id][0],
                                    'amount_tva': tax_dict[tax_line.tax_id.id][1],
                                    'amount_ttc': tax_dict[tax_line.tax_id.id][2],
                                    'invoice_date': move_pay.invoice_id.date_invoice,
                                    'paiement_date': move.date,
                                    'paiement_type': pay_method,
                                    'tax_rate': round(tax_line.tax_id.amount,2),
                                    'invoice_tax_id': tax_line.id,
                                    'tax_id': tax_line.tax_id.id,
                                    'type': type
                                }
                                tva_declaration_line_obj.create(data)
        self.get_tva_credit()


class TvaDeclarationLine(models.Model):
    _name = "tva.declaration.line"
    _description = "Ligne Declaration Tva"
    _order = 'sequence, id'

    tva_declaration_id = fields.Many2one('tva.declaration', u'Déclaration TVA', ondelete='cascade')
    sequence = fields.Integer(string=u'Séquence')
    invoice_id = fields.Many2one('account.invoice', 'Facture', readonly=False)
    invoice_number = fields.Char(u'N°Facture')
    description = fields.Char('Description')
    amount_ht = fields.Float('Montant HT')
    amount_tva = fields.Float('Montant TVA')
    amount_ttc = fields.Float('Montant TTC')
    invoice_date = fields.Date('Date facture')
    paiement_date = fields.Date('Date de paiement')
    paiement_type = fields.Many2one(comodel_name='payement.method',
                                         string=u'Méthode de paiement', readonly=False)
    tax_rate = fields.Float('Taux de TVA', digits=(16,2))
    invoice_tax_id = fields.Many2one('account.invoice.tax', 'Taxe facture')
    tax_id = fields.Many2one('account.tax', 'TVA')
    type = fields.Boolean('Encaissement?')
    partner_id = fields.Many2one('res.partner', 'Partenaire')
    id_fisc = fields.Char('IF')
    partner_name = fields.Char('Nom FR')
    ice = fields.Char('ICE')
    to_be_delayed = fields.Boolean('To be delayed?')
    delayed = fields.Boolean('Delayed ?')


class TvaLine(models.Model):
    _name = "tva.line"
    _description = "Lignes Tva"

    tva_declaration_id = fields.Many2one('tva.declaration', u'Déclaration TVA', ondelete='cascade')
    tax_id = fields.Many2one('account.tax', 'TVA')
    amount_ht = fields.Float('Base TVA')
    amount = fields.Float('Montant TVA')
    type = fields.Selection([
        ('r', 'Fournisseur'),
        ('p', 'Client')])